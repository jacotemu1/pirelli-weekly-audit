from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Finding, PageResult

SEVERITY_RANK = {'Critica': 4, 'Alta': 3, 'Media': 2, 'Bassa': 1}
SEVERITY_WEIGHT = {'Critica': 10, 'Alta': 5, 'Media': 2, 'Bassa': 1}
CATEGORY_SHEETS = {
    'technical': '11_Bug_Codice_Tecnica',
    'seo': '12_Bug_SEO',
    'ux': '13_Bug_UX_UI',
    'content': '14_Bug_Contenuti_Localizzazione',
    'accessibility': '15_Bug_Accessibilita',
    'cro': '16_Bug_CRO',
}


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or '')))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def _sheet_from_df(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(name)
    if df.empty:
        ws['A1'] = 'No data'
        return
    for row_idx, row in enumerate([df.columns.tolist()] + df.values.tolist(), start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if row_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill('solid', fgColor='D9EAD3')
    _auto_width(ws)


def _finding_confidence(f: Finding) -> str:
    if isinstance(f.data, dict):
        return str(f.data.get('confidence') or 'Media')
    return 'Media'


def _finding_evidence(f: Finding) -> str:
    if isinstance(f.data, dict) and f.data.get('evidenza_tecnica'):
        return str(f.data['evidenza_tecnica'])
    return f'category={f.category}; severity={f.severity}; url={f.url}'


def _finding_discovered_from(f: Finding) -> str:
    if isinstance(f.data, dict):
        return str(f.data.get('discovered_from') or '')
    return ''


def _finding_crawl_depth(f: Finding) -> int:
    if isinstance(f.data, dict):
        try:
            return int(f.data.get('crawl_depth') or 0)
        except Exception:
            return 0
    return 0


def build_excel(output_path: str | Path, pages: list[PageResult], findings: list[Finding], diff: dict[str, set[str]], run_date: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    pages_df = pd.DataFrame([
        {
            'site_code': p.site_code,
            'country': p.country,
            'region': p.region,
            'language': p.language,
            'page_type': p.page_type,
            'url': p.url,
            'final_url': p.final_url,
            'status_code': p.status_code,
            'title': p.title,
            'h1': p.h1,
            'h2_count': p.h2_count,
            'canonical': p.canonical,
            'meta_description': p.meta_description,
            'crawl_depth': getattr(p, 'crawl_depth', 0),
            'discovered_from': getattr(p, 'discovered_from', ''),
            'errors': ' | '.join(p.errors),
        }
        for p in pages
    ])
    bugs_df = pd.DataFrame([
        {
            'mercato': f.site_code,
            'paese': f.country,
            'regione': f.region,
            'area_bug': f.category,
            'gravita': f.severity,
            'titolo_bug': f.title,
            'spiegazione_bug_it': f.description,
            'impatto_utenti_business': f.impact,
            'fix_consigliato_it': f.suggested_fix,
            'evidenza_tecnica': _finding_evidence(f),
            'confidence': _finding_confidence(f),
            'pagina': f.url,
            'tipo_pagina': f.page_type,
            'pagina_trovata_da': _finding_discovered_from(f),
            'crawl_depth': _finding_crawl_depth(f),
            'fingerprint': f.fingerprint,
            'stato_settimanale': 'new' if f.fingerprint in diff['new'] else 'persistent' if f.fingerprint in diff['persistent'] else '',
        }
        for f in findings
    ])

    counts_by_site = Counter(f.site_code for f in findings)
    crit_by_site = Counter(f.site_code for f in findings if f.severity == 'Critica')
    high_by_site = Counter(f.site_code for f in findings if f.severity == 'Alta')
    site_pages = Counter(p.site_code for p in pages)
    site_regions = {p.site_code: p.region for p in pages}
    site_countries = {p.site_code: p.country for p in pages}
    coverage_rows = []
    for site_code in sorted(site_pages):
        total_pages = site_pages[site_code]
        total_findings = counts_by_site.get(site_code, 0)
        priority_score = sum(SEVERITY_WEIGHT.get(f.severity, 1) for f in findings if f.site_code == site_code)
        quality_score = max(0, 100 - (crit_by_site[site_code] * 25 + high_by_site[site_code] * 12 + max(total_findings - crit_by_site[site_code] - high_by_site[site_code], 0) * 4))
        coverage_rows.append(
            {
                'mercato': site_code,
                'paese': site_countries.get(site_code, ''),
                'regione': site_regions.get(site_code, ''),
                'pagine_crawlate': total_pages,
                'bug_totali': total_findings,
                'bug_critici': crit_by_site.get(site_code, 0),
                'bug_alti': high_by_site.get(site_code, 0),
                'priority_score': priority_score,
                'quality_score_estimate': quality_score,
            }
        )
    coverage_df = pd.DataFrame(coverage_rows)

    sintesi_df = pd.DataFrame(
        [
            {'metrica': 'run_date', 'valore': run_date},
            {'metrica': 'pagine_crawlate', 'valore': len(pages)},
            {'metrica': 'bug_totali', 'valore': len(findings)},
            {'metrica': 'bug_nuovi', 'valore': len(diff['new'])},
            {'metrica': 'bug_risolti', 'valore': len(diff['resolved'])},
            {'metrica': 'bug_persistenti', 'valore': len(diff['persistent'])},
        ]
    )

    priorita_df = coverage_df.sort_values(['priority_score', 'bug_totali'], ascending=[False, False]) if not coverage_df.empty else pd.DataFrame()
    diff_df = bugs_df[['mercato', 'paese', 'area_bug', 'gravita', 'titolo_bug', 'pagina', 'fingerprint', 'stato_settimanale']].copy() if not bugs_df.empty else pd.DataFrame()
    pages_sheet_df = pages_df.rename(
        columns={
            'site_code': 'mercato',
            'country': 'paese',
            'region': 'regione',
            'language': 'lingua',
            'page_type': 'tipo_pagina',
            'url': 'url_seed',
        }
    )

    raw_rows = []
    for p in pages:
        raw_rows.append(
            {
                'mercato': p.site_code,
                'url': p.final_url or p.url,
                'status_code': p.status_code,
                'title': p.title,
                'h1': p.h1,
                'meta_description': p.meta_description,
                'canonical': p.canonical,
                'links_count': len(p.links),
                'html_chars': len(p.html or ''),
                'text_chars': len(p.text or ''),
                'errors': ' | '.join(p.errors),
            }
        )
    raw_df = pd.DataFrame(raw_rows)

    _sheet_from_df(wb, '00_Sintesi', sintesi_df)
    _sheet_from_df(wb, '01_Priorita', priorita_df)
    _sheet_from_df(wb, '02_Diff_settimanale', diff_df)
    _sheet_from_df(wb, '10_Bug_Tutti', bugs_df)
    for category, sheet_name in CATEGORY_SHEETS.items():
        category_df = bugs_df[bugs_df['area_bug'] == category].copy() if not bugs_df.empty else pd.DataFrame()
        _sheet_from_df(wb, sheet_name, category_df)
    _sheet_from_df(wb, '90_Pagine_Crawlate', pages_sheet_df)
    _sheet_from_df(wb, '91_Coverage', coverage_df)
    _sheet_from_df(wb, '99_Raw_Tecnico', raw_df)

    build_info_df = pd.DataFrame(
        [
            {'chiave': 'build_version', 'valore': 'V4_20260331'},
            {'chiave': 'generated_at_utc', 'valore': datetime.utcnow().isoformat(timespec='seconds')},
            {'chiave': 'run_date', 'valore': run_date},
            {'chiave': 'pages_checked', 'valore': len(pages)},
            {'chiave': 'findings', 'valore': len(findings)},
        ]
    )
    _sheet_from_df(wb, 'Build Info', build_info_df)

    wb.save(output_path)


def build_markdown_summary(output_path: str | Path, pages: list[PageResult], findings: list[Finding], diff: dict[str, set[str]], run_date: str) -> None:
    by_site: dict[str, list[Finding]] = defaultdict(list)
    country_map = {}
    for f in findings:
        by_site[f.site_code].append(f)
        country_map[f.site_code] = f.country
    for p in pages:
        country_map[p.site_code] = p.country

    lines = [
        '# Pirelli Weekly Audit Summary',
        '',
        f'- Run date: {run_date}',
        f'- Pages checked: {len(pages)}',
        f'- Findings total: {len(findings)}',
        f'- New findings: {len(diff["new"])}',
        f'- Resolved findings: {len(diff["resolved"])}',
        '',
        '## Highlights',
        '',
    ]
    if findings:
        severity_sorted = sorted(findings, key=lambda f: (-SEVERITY_RANK.get(f.severity, 0), f.country, f.title))[:10]
        for f in severity_sorted:
            lines.append(f'- **{f.country}** — {f.severity} — {f.title}: {f.description}')
    else:
        lines.append('- Nessuna issue rilevata.')

    lines.extend(['', '## Per market', ''])
    for site_code in sorted(country_map):
        country = country_map[site_code]
        site_findings = by_site.get(site_code, [])
        counter = Counter(f.severity for f in site_findings)
        lines.append(f'### {country} ({site_code})')
        lines.append(f'- Findings: {len(site_findings)}')
        lines.append(f'- Critiche: {counter.get("Critica", 0)} | Alte: {counter.get("Alta", 0)} | Medie: {counter.get("Media", 0)} | Basse: {counter.get("Bassa", 0)}')
        for f in sorted(site_findings, key=lambda x: -SEVERITY_RANK.get(x.severity, 0))[:5]:
            lines.append(f'  - {f.severity} — {f.title}: {f.description}')
        lines.append('')

    Path(output_path).write_text('\n'.join(lines), encoding='utf-8')
