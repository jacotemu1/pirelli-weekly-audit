from __future__ import annotations

from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Finding, PageResult

BUILD_VERSION = 'V4_20260331'
SEVERITY_WEIGHT = {'Critica': 10, 'Alta': 5, 'Media': 2, 'Bassa': 1}
SEVERITY_SCORE = {'Critica': 4, 'Alta': 3, 'Media': 2, 'Bassa': 1}
CONFIDENCE_SCORE = {'Alta': 3, 'Media': 2, 'Bassa': 1}
CATEGORY_SHEETS = {
    'Codice/Tecnica': '11_Bug_Codice_Tecnica',
    'SEO': '12_Bug_SEO',
    'UX/UI': '13_Bug_UX_UI',
    'Contenuti/Localizzazione': '14_Bug_Contenuti_Localizzazione',
    'Accessibilità': '15_Bug_Accessibilita',
    'CRO': '16_Bug_CRO',
}
SHEET_COLORS = {
    'summary': 'D9EAD3',
    'priority': 'FCE5CD',
    'diff': 'D9EAD3',
    'bugs': 'FFF2CC',
    'raw': 'D9EAD3',
}



def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or '')))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 46)



def _apply_header_style(ws, row_num: int = 1, fill: str = 'D9EAD3') -> None:
    for cell in ws[row_num]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor=fill)
        cell.alignment = Alignment(vertical='top', wrap_text=True)



def _sheet_from_df(wb: Workbook, name: str, df: pd.DataFrame, fill: str = 'D9EAD3') -> None:
    ws = wb.create_sheet(name)
    if df.empty:
        ws['A1'] = 'No data'
    else:
        for row_idx, row in enumerate([df.columns.tolist()] + df.values.tolist(), start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        _apply_header_style(ws, 1, fill)
    _auto_width(ws)



def build_excel(output_path: str | Path, pages: list[PageResult], findings: list[Finding], diff: dict[str, set[str]], run_date: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    pages_df = pd.DataFrame([
        {
            'mercato': p.site_code,
            'paese': p.country,
            'regione': p.region,
            'lingua': p.language,
            'tipo_pagina': p.page_type,
            'crawl_depth': p.crawl_depth,
            'discovered_from': p.discovered_from,
            'url': p.url,
            'final_url': p.final_url,
            'status_code': p.status_code,
            'title': p.title,
            'h1': p.h1,
            'h2_count': p.h2_count,
            'canonical': p.canonical,
            'meta_description': p.meta_description,
            'out_links_count': len(p.links),
            'errors': ' | '.join(p.errors),
        }
        for p in pages
    ])

    findings_rows = []
    for f in findings:
        diff_status = 'new' if f.fingerprint in diff['new'] else 'persistent' if f.fingerprint in diff['persistent'] else ''
        findings_rows.append({
            'mercato': f.site_code,
            'paese': f.country,
            'regione': f.region,
            'area_bug': f.category,
            'gravita': f.severity,
            'confidence': f.confidence,
            'titolo_bug': f.title,
            'spiegazione_bug_it': f.explanation_it,
            'impatto_utenti_business': f.impact_it,
            'fix_consigliato_it': f.suggested_fix_it,
            'evidenza_tecnica': f.evidence_tecnica,
            'pagina': f.url,
            'tipo_pagina': f.page_type,
            'pagina_trovata_da': f.discovered_from,
            'crawl_depth': f.crawl_depth,
            'stato_settimanale': diff_status,
            'fingerprint': f.fingerprint,
            '_severity_score': SEVERITY_SCORE.get(f.severity, 0),
            '_confidence_score': CONFIDENCE_SCORE.get(f.confidence, 0),
            '_page_priority': 3 if f.page_type == 'dealer' else 2 if f.page_type in {'catalogue', 'home'} else 1,
        })
    findings_df = pd.DataFrame(findings_rows)

    site_pages = Counter(p.site_code for p in pages)
    site_regions = {p.site_code: p.region for p in pages}
    site_countries = {p.site_code: p.country for p in pages}
    by_site_findings = Counter(f.site_code for f in findings)
    by_site_200 = Counter(p.site_code for p in pages if (p.status_code or 0) < 400)
    by_site_non200 = Counter(p.site_code for p in pages if p.status_code is None or p.status_code >= 400)
    max_depth = Counter()
    severity_penalty = Counter()
    for p in pages:
        max_depth[p.site_code] = max(max_depth.get(p.site_code, 0), p.crawl_depth)
    for f in findings:
        severity_penalty[f.site_code] += SEVERITY_WEIGHT.get(f.severity, 0)

    summary_rows = []
    for site_code in sorted(site_pages):
        summary_rows.append({
            'mercato': site_code,
            'paese': site_countries.get(site_code, ''),
            'regione': site_regions.get(site_code, ''),
            'pagine_analizzate': site_pages[site_code],
            'pagine_ok': by_site_200[site_code],
            'pagine_non_ok': by_site_non200[site_code],
            'profondita_massima': max_depth[site_code],
            'bug_totali': by_site_findings[site_code],
            'score_qualita_stimato': max(0, 100 - severity_penalty[site_code]),
        })
    summary_df = pd.DataFrame(summary_rows)
    if not summary_df.empty:
        summary_df = summary_df.sort_values(['score_qualita_stimato', 'bug_totali'], ascending=[False, True])

    # 00_Sintesi
    ws = wb.create_sheet('00_Sintesi')
    ws['A1'] = f'Pirelli Weekly Audit {BUILD_VERSION}'
    ws['A1'].font = Font(bold=True, size=14)
    metrics = [
        ('Build version', BUILD_VERSION),
        ('Run date', run_date),
        ('Pagine analizzate', len(pages)),
        ('Bug totali', len(findings)),
        ('Bug critici', sum(1 for f in findings if f.severity == 'Critica')),
        ('Bug alti', sum(1 for f in findings if f.severity == 'Alta')),
        ('Nuovi bug', len(diff['new'])),
        ('Bug risolti', len(diff['resolved'])),
        ('Bug persistenti', len(diff['persistent'])),
    ]
    for idx, (k, v) in enumerate(metrics, start=3):
        ws.cell(row=idx, column=1, value=k)
        ws.cell(row=idx, column=2, value=v)

    ws['D3'] = 'Top 5 mercati con più bug'
    ws['D3'].font = Font(bold=True)
    for i, row in enumerate(sorted(summary_rows, key=lambda r: (-r['bug_totali'], r['score_qualita_stimato']))[:5], start=4):
        ws.cell(row=i, column=4, value=row['paese'])
        ws.cell(row=i, column=5, value=row['bug_totali'])

    ws['G3'] = 'Top 5 problemi globali'
    ws['G3'].font = Font(bold=True)
    title_counter = Counter(f.title for f in findings)
    for i, (title, count) in enumerate(title_counter.most_common(5), start=4):
        ws.cell(row=i, column=7, value=title)
        ws.cell(row=i, column=8, value=count)

    ws['J3'] = 'Quick wins'
    ws['J3'].font = Font(bold=True)
    quick_wins = [
        'Correggere placeholder e stringhe non localizzate nel dealer finder.',
        'Ridurre CTA duplicate nelle homepage dei mercati più rumorosi.',
        'Verificare 404 reali su pagine linkate internamente.',
        'Pulire contenuti con anno passato o promo scadute.',
        'Controllare heading/title mancanti o duplicati nei template shared.',
    ]
    for i, item in enumerate(quick_wins, start=4):
        ws.cell(row=i, column=10, value=item)
    _auto_width(ws)

    # 01_Priorita
    if findings_df.empty:
        priority_df = findings_df
    else:
        priority_df = findings_df.sort_values(
            ['_severity_score', '_confidence_score', '_page_priority', 'crawl_depth', 'mercato'],
            ascending=[False, False, False, True, True],
        ).drop(columns=['_severity_score', '_confidence_score', '_page_priority'])
    _sheet_from_df(wb, '01_Priorita', priority_df, SHEET_COLORS['priority'])

    # 02_Diff_settimanale
    diff_rows = [
        {'tipo': 'new', 'conteggio': len(diff['new'])},
        {'tipo': 'resolved', 'conteggio': len(diff['resolved'])},
        {'tipo': 'persistent', 'conteggio': len(diff['persistent'])},
    ]
    _sheet_from_df(wb, '02_Diff_settimanale', pd.DataFrame(diff_rows), SHEET_COLORS['diff'])

    # 10 total bug sheet + category sheets
    bug_public_cols = [
        'mercato', 'paese', 'regione', 'area_bug', 'gravita', 'confidence', 'titolo_bug', 'spiegazione_bug_it',
        'impatto_utenti_business', 'fix_consigliato_it', 'evidenza_tecnica', 'pagina', 'tipo_pagina',
        'pagina_trovata_da', 'crawl_depth', 'stato_settimanale', 'fingerprint'
    ]
    bug_df_public = findings_df[bug_public_cols] if not findings_df.empty else pd.DataFrame(columns=bug_public_cols)
    _sheet_from_df(wb, '10_Bug_Tutti', bug_df_public, SHEET_COLORS['bugs'])

    if not findings_df.empty:
        for category, sheet_name in CATEGORY_SHEETS.items():
            subset = findings_df[findings_df['area_bug'] == category][bug_public_cols]
            _sheet_from_df(wb, sheet_name, subset, SHEET_COLORS['bugs'])
    else:
        for sheet_name in CATEGORY_SHEETS.values():
            _sheet_from_df(wb, sheet_name, pd.DataFrame(columns=bug_public_cols), SHEET_COLORS['bugs'])

    # Raw / coverage sheets
    coverage_rows = []
    for site_code in sorted(site_pages):
        sample_urls = [p.final_url or p.url for p in pages if p.site_code == site_code][:5]
        coverage_rows.append({
            'mercato': site_code,
            'paese': site_countries.get(site_code, ''),
            'pagine_analizzate': site_pages[site_code],
            'profondita_massima': max_depth[site_code],
            'sample_urls': ' | '.join(sample_urls),
        })
    coverage_df = pd.DataFrame(coverage_rows)
    _sheet_from_df(wb, '90_Pagine_Crawlate', pages_df[['mercato', 'paese', 'regione', 'tipo_pagina', 'crawl_depth', 'discovered_from', 'url', 'final_url', 'status_code', 'title', 'h1']], SHEET_COLORS['raw']) if not pages_df.empty else _sheet_from_df(wb, '90_Pagine_Crawlate', pages_df, SHEET_COLORS['raw'])
    _sheet_from_df(wb, '91_Coverage', coverage_df, SHEET_COLORS['raw'])
    _sheet_from_df(wb, '99_Raw_Tecnico', pages_df, SHEET_COLORS['raw'])

    build_df = pd.DataFrame([
        {'key': 'build_version', 'value': BUILD_VERSION},
        {'key': 'run_date', 'value': run_date},
        {'key': 'pages_checked', 'value': len(pages)},
        {'key': 'findings_total', 'value': len(findings)},
    ])
    _sheet_from_df(wb, 'Build Info', build_df, SHEET_COLORS['raw'])

    wb.save(output_path)



def build_markdown_summary(output_path: str | Path, pages: list[PageResult], findings: list[Finding], diff: dict[str, set[str]], run_date: str) -> None:
    by_site = Counter(f.site_code for f in findings)
    site_country = {p.site_code: p.country for p in pages}
    lines = [
        '# Pirelli Weekly Audit Summary',
        '',
        f'- Build version: {BUILD_VERSION}',
        f'- Run date: {run_date}',
        f'- Pages checked: {len(pages)}',
        f'- Findings total: {len(findings)}',
        f'- New findings: {len(diff["new"])}',
        f'- Resolved findings: {len(diff["resolved"])}',
        f'- Persistent findings: {len(diff["persistent"])}',
        '',
        '## Top mercati per numero di bug',
        '',
    ]
    for site_code, count in by_site.most_common(10):
        lines.append(f'- {site_country.get(site_code, site_code)} ({site_code}): {count} bug')
    Path(output_path).write_text('\n'.join(lines), encoding='utf-8')
