from __future__ import annotations

from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Finding, PageResult

BUILD_VERSION = 'V3_FIXED_20260331'
SEVERITY_WEIGHT = {'Critica': 10, 'Alta': 5, 'Media': 2, 'Bassa': 1}


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or '')))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 42)


def build_excel(output_path: str | Path, pages: list[PageResult], findings: list[Finding], diff: dict[str, set[str]], run_date: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    pages_df = pd.DataFrame([
        {
            'site_code': p.site_code,
            'country': p.country,
            'region': p.region,
            'language': p.language,
            'page_type': p.page_type,
            'crawl_depth': p.crawl_depth,
            'discovered_from': p.discovered_from,
            'url': p.url,
            'final_url': p.final_url,
            'status_code': p.status_code,
            'title': p.title,
            'h1': p.h1,
            'h2_count': p.h2_count,
            'canonical': p.canonical,
            'meta_description': p.meta_description,
            'out_links_count': len(p.links),
            'errors': ' | '.join(p.errors),
        }
        for p in pages
    ])
    findings_df = pd.DataFrame([
        {
            'site_code': f.site_code,
            'country': f.country,
            'region': f.region,
            'page_type': f.page_type,
            'crawl_depth': f.crawl_depth,
            'pagina_trovata_da': f.discovered_from,
            'url': f.url,
            'area': f.category,
            'severity': f.severity,
            'confidence': f.confidence,
            'titolo_bug': f.title,
            'spiegazione_bug_it': f.explanation_it,
            'impatto_utenti_business': f.impact_it,
            'fix_consigliato_it': f.suggested_fix_it,
            'evidenza_tecnica': f.evidence_tecnica,
            'fingerprint': f.fingerprint,
            'diff_status': 'new' if f.fingerprint in diff['new'] else 'persistent' if f.fingerprint in diff['persistent'] else '',
        }
        for f in findings
    ])

    site_pages = Counter(p.site_code for p in pages)
    by_site_findings = Counter(f.site_code for f in findings)
    by_site_200 = Counter(p.site_code for p in pages if (p.status_code or 0) < 400)
    by_site_non200 = Counter(p.site_code for p in pages if p.status_code is None or p.status_code >= 400)
    max_depth = Counter()
    site_regions = {}
    site_countries = {}
    severity_penalty = Counter()
    for p in pages:
        site_regions[p.site_code] = p.region
        site_countries[p.site_code] = p.country
        max_depth[p.site_code] = max(max_depth.get(p.site_code, 0), p.crawl_depth)
    for f in findings:
        severity_penalty[f.site_code] += SEVERITY_WEIGHT.get(f.severity, 0)

    summary_rows = []
    for site_code in sorted(site_pages):
        score = max(0, 100 - severity_penalty[site_code])
        summary_rows.append({
            'site_code': site_code,
            'country': site_countries.get(site_code, ''),
            'region': site_regions.get(site_code, ''),
            'pages_checked': site_pages[site_code],
            'pages_200': by_site_200[site_code],
            'pages_non_200': by_site_non200[site_code],
            'max_crawl_depth': max_depth[site_code],
            'findings_total': by_site_findings[site_code],
            'quality_score_estimate': score,
        })
    summary_df = pd.DataFrame(summary_rows).sort_values(['quality_score_estimate', 'findings_total'], ascending=[False, True]) if summary_rows else pd.DataFrame()

    summary_ws = wb.create_sheet('Summary')
    summary_ws['A1'] = f'Pirelli Weekly Audit {BUILD_VERSION}'
    summary_ws['A1'].font = Font(bold=True, size=14)
    summary_ws['A2'] = 'Build version'
    summary_ws['B2'] = BUILD_VERSION
    summary_ws['A3'] = 'Run date'
    summary_ws['B3'] = run_date
    summary_ws['A4'] = 'Pages checked'
    summary_ws['B4'] = len(pages)
    summary_ws['A5'] = 'Findings total'
    summary_ws['B5'] = len(findings)
    summary_ws['A6'] = 'New findings vs previous run'
    summary_ws['B6'] = len(diff['new'])
    summary_ws['A7'] = 'Resolved findings vs previous run'
    summary_ws['B7'] = len(diff['resolved'])
    summary_ws['A8'] = 'Persistent findings vs previous run'
    summary_ws['B8'] = len(diff['persistent'])
    summary_ws['A10'] = 'Per-site summary'
    summary_ws['A10'].font = Font(bold=True)
    if not summary_df.empty:
        for row_idx, row in enumerate([summary_df.columns.tolist()] + summary_df.values.tolist(), start=11):
            for col_idx, value in enumerate(row, start=1):
                cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if row_idx == 11:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill('solid', fgColor='D9EAD3')
    _auto_width(summary_ws)

    coverage_rows = []
    for site_code in sorted(site_pages):
        sample_urls = [p.final_url or p.url for p in pages if p.site_code == site_code][:5]
        coverage_rows.append({
            'site_code': site_code,
            'country': site_countries.get(site_code, ''),
            'pages_checked': site_pages[site_code],
            'max_crawl_depth': max_depth[site_code],
            'sample_urls': ' | '.join(sample_urls),
        })
    coverage_df = pd.DataFrame(coverage_rows)

    build_ws = wb.create_sheet('Build Info')
    build_ws.append(['key', 'value'])
    build_ws.append(['build_version', BUILD_VERSION])
    build_ws.append(['run_date', run_date])
    build_ws.append(['pages_checked', len(pages)])
    build_ws.append(['findings_total', len(findings)])
    for cell in build_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='D9EAD3')
    _auto_width(build_ws)

    for name, df in [('Pages', pages_df), ('Findings', findings_df), ('Coverage', coverage_df)]:
        ws = wb.create_sheet(name)
        if df.empty:
            ws['A1'] = 'No data'
        else:
            for row_idx, row in enumerate([df.columns.tolist()] + df.values.tolist(), start=1):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    if row_idx == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill('solid', fgColor='D9EAD3')
        _auto_width(ws)

    diff_ws = wb.create_sheet('Weekly Diff')
    diff_ws.append(['type', 'count'])
    diff_ws.append(['new', len(diff['new'])])
    diff_ws.append(['resolved', len(diff['resolved'])])
    diff_ws.append(['persistent', len(diff['persistent'])])
    for cell in diff_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='D9EAD3')
    _auto_width(diff_ws)

    wb.save(output_path)


def build_markdown_summary(output_path: str | Path, pages: list[PageResult], findings: list[Finding], diff: dict[str, set[str]], run_date: str) -> None:
    by_site = Counter(f.site_code for f in findings)
    site_country = {p.site_code: p.country for p in pages}
    lines = [
        '# Pirelli Weekly Audit Summary',
        '',
        f'- Run date: {run_date}',
        f'- Pages checked: {len(pages)}',
        f'- Findings total: {len(findings)}',
        f'- New findings: {len(diff["new"])}',
        f'- Resolved findings: {len(diff["resolved"])}',
        f'- Persistent findings: {len(diff["persistent"])}',
        '',
        '## Top mercati per numero di findings',
        '',
    ]
    for site_code, count in by_site.most_common(10):
        lines.append(f'- {site_country.get(site_code, site_code)} ({site_code}): {count} findings')
    Path(output_path).write_text('\n'.join(lines), encoding='utf-8')
